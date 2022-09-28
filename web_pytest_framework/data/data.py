# -*- codeing = utf-8 -*-
# @Time :2022/9/23 9:22
# @Author :ctq
# @File : data.py.py
# @Desc :
from study.web自动化测试.config.config import  data_path
class ReadWrite:
    def __init__(self):
        self.txtpath=r''
        self.excelpath=data_path
        self.yamlpath=r''


    def txtread(self):
        list1=[]
        with open(self.txtpath,'r',encoding='utf-8') as f:
            values=f.readlines()
            print(values)
            for data in values:
                data_v=data.strip('\n')
                list1.append(data_v)
            return list1
    def txtwrite(self,username,password):
        with open(self.txtpath,'r',encoding='utf-8') as f1:
            old_f=f1.read()
            with open(self.file,'a',encoding='utf-8') as f2:
                while True:
                    if self.username in old_f:
                        self.username=input("用户名已存在，请重新输入:")
                        self.password=input("请输入密码:")
                    else:
                        value=f"{self.username},{self.password}\n"
                        f2.writelines(value)
                        print("注册成功!")
                        return True

    def excelread(self,sheetname):
        import openpyxl
        wb=openpyxl.load_workbook(self.excelpath)
        table=wb[sheetname]
        list1=[]
        for row in range(2,table.max_row+1):
            list2=[]
            for col in range(1,table.max_column+1):
                values=table.cell(row,col).value
                list2.append(values)
            list1.append(list2)
        return  list1
    def excelwrite(self,*args,sheetname): #位置参数要放在可变参数后面
        import  openpyxl
        wb=openpyxl.load_workbook(self.excelpath)
        table=wb[sheetname]
        row=table.max_row
        col=len(args)
        for col in range(col):
            table.cell(row+1,col+1).value=args[col]
        wb.save(self.excelpath)
    def mysqlread(self):
        import pymysql
        db=pymysql.connect(host='localhost',port=3306,user='root',password='root',database='T31',charset='utf8')
        cur=db.cursor()
        sql='select * from users where username="test1"'
        cur.execute(sql)
        content=cur.fetchall()
        return  content

    def mysqlwrite(self,username,password):
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='root', database='T31', charset='utf8')
        cur = db.cursor()
        sql = f'insert into users values ("{username}","{password}")  '
        cur.execute(sql)
        db.commit()

    def yamlread(self):
        import  yaml
        f = open(self.yamlpath, 'r', encoding='utf-8')
        content = f.read()
        data = yaml.safe_load(content)  # 转化成字典模式
        return  data

    def yamlwrite(self,username,password):
        import yaml
        f = open(yamlpath, 'a', encoding='utf-8')
        data={'username':username,'password':password}
        yaml.safe_dump(data,f)
        f.close()


if __name__=="__main__":
    t1=ReadWrite()
    print(t1.excelread('users'))
    pass